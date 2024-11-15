import os.path
import warnings
from abc import ABC, abstractmethod
from datetime import datetime

import pandas as pd
from openpyxl import load_workbook
from openpyxl.chart import Reference
from openpyxl.chart import ScatterChart, Series
from openpyxl.utils import column_index_from_string, get_column_letter

from src.modules.project.projects.project_interface import ProjectInterface
from src.utilities.in_out_files import write_pd_to_csv


class ProjectOutputWriterInterface(ABC):

    @property
    @abstractmethod
    def project(self) -> ProjectInterface:
        pass

    @project.setter
    @abstractmethod
    def project(self, value: ProjectInterface):
        pass

    @property
    @abstractmethod
    def out_dir_path(self):
        pass

    @out_dir_path.setter
    @abstractmethod
    def out_dir_path(self):
        pass

    @abstractmethod
    def dump(self):
        pass


class LAMMPSProjectOutputWriterInterface(ProjectOutputWriterInterface):
    def __init__(self, project: ProjectInterface):
        self._project = project
        self._out_dir_path = os.path.join(project.path_to_proj_dir, project.configs["DirectoryPaths"]["REL_PATH_OUTPUT_DIR"])
        os.makedirs(self._out_dir_path) if not os.path.exists(self._out_dir_path) else ...

    @property
    def project(self) -> ProjectInterface:
        return self._project

    @project.setter
    def project(self, value: ProjectInterface):
        self._project = value

    @property
    def out_dir_path(self):
        return self._out_dir_path

    @out_dir_path.setter
    def out_dir_path(self, value):
        self._out_dir_path = value

    @staticmethod
    def contains_other_data_structures(df: pd.DataFrame) -> bool:
        for row in df.itertuples(index=False):
            for cell in row:
                if not isinstance(cell, pd.DataFrame):
                    return True
        return False


class LAMMPSProjectScalerGlobalOutputWriter(LAMMPSProjectOutputWriterInterface):

    def __init__(self, project: ProjectInterface):
        super().__init__(project)

    def dump(self):
        # remove_folder_contents(self._out_dir_path)
        for data_name, data in self.project.project_outputs.items():
            if isinstance(data.df, pd.DataFrame):
                # if not self.contains_other_data_structures(data.df):
                date_time_str = datetime.now().strftime("%Y%m%d_%H%M%S")
                file_name = f'{data_name}_{date_time_str}.csv'
                write_pd_to_csv(self._out_dir_path, file_name, data.df)
                # else:
                #     warnings.warn("No global scalar property was found in the project. Nothing dumped!")


class LAMMPSProjectProfileOutputWriter(LAMMPSProjectOutputWriterInterface):
    def __init__(self, project: ProjectInterface):
        super().__init__(project)

    def dump(self):
        for data_name, data in self.project.project_outputs.items():
            df_data = data.df
            if isinstance(df_data, pd.DataFrame):
                if self.contains_other_data_structures(df_data):  # Each record in the dataframe contains another data structure.
                    date_time_str = datetime.now().strftime("%Y%m%d_%H%M%S")
                    file_name = f'{data_name}_{date_time_str}.xlsx'
                    excel_file_path = os.path.join(self._out_dir_path, file_name)
                    with pd.ExcelWriter(excel_file_path, engine='openpyxl') as df_to_excel_writer:
                        # Need to create an excel-sheet that entails the experimental data
                        for row in df_data.itertuples(index=False):
                            for cell in row:
                                if isinstance(cell, list):  # List items belong to the same property.
                                    start_column = 0
                                    for list_item in cell:
                                        if not isinstance(list_item, dict):
                                            continue
                                        for key, value in list_item.items():
                                            if key == 'profile':
                                                profile_df = value[0]['data']
                                                profile_df.to_excel(df_to_excel_writer, sheet_name=row.ff_filename, index=False,
                                                                    startcol=start_column)
                                                # Calculate the starting column for the next DataFrame (len(df1.columns) + a buffer)
                                                start_column = len(profile_df.columns) + 2  # For example, 2-column buffer between DataFrames
                else:
                    warnings.warn("No profile was found in the project. Nothing dumped!")

                # Now open the written excel file to append user-specific data
                exp_data = {
                    'A1': 'T, K',
                    'B1': 274.83,
                    'A2': 'p, atm',
                    'B2': 11.85,
                    'A3': 'x_co2, mole/mole',
                    'B3': 0.0142,
                    'D1': 'mw_guest (gr/mol)',
                    'E1': 44.01,
                    'D2': 'mw_host (gr/mol)',
                    'E2': 18
                }

                self.write_to_excel(excel_file_path, 'sim_settings', exp_data)

                data_to_write = {
                    'E1': 'Average Density (gcc)',
                    'F1': 'Average Density (mol/cc)',
                    'K1': 'Average Density (gcc)',
                    'L1': 'Average Density (mol/cc)',
                    'M1': 'x_guest (mole/mole)',
                    'N1': 'RE (%)',
                }
                formulas_to_write = {
                    'E2': 'AVERAGE(D2:D20)',
                    'F2': 'E2/sim_settings!E1',
                    'K2': 'AVERAGE(J2:J20)',
                    'L2': 'K2/sim_settings!E2',
                    'M2': 'F2/(F2+L2)',
                    'N2': '100*(M2-sim_settings!B3)/sim_settings!B3'

                }

                sheets_to_exclude = ['sim_settings']  # Replace with actual sheet names you want to exclude
                wb = load_workbook(excel_file_path)
                # Loop through all sheets in the workbook and apply the function, skipping the ones to exclude
                for sheet_name in wb.sheetnames:
                    if sheet_name not in sheets_to_exclude:
                        self.write_to_excel(excel_file_path, sheet_name, data_to_write, formulas_to_write)
                        # Add chart to the excel sheet.
                        self.add_multiple_scatter_charts_to_sheet(excel_file_path, sheet_name, [2, 8], [4, 10], [2, 2], [22, 22], [2, 2], [22, 22])
                        # self.add_scatter_chart_to_sheet(excel_file_path, sheet_name, 2, 4, 2, 22, 2, 22)
                        # self.add_scatter_chart_to_sheet(excel_file_path, sheet_name, 8, 10, 2, 22, 2, 22)

    @staticmethod
    def write_to_excel(file_path, sheet_name, data_to_write=None, formulas_to_write=None):
        """
        Writes data and formulas to the specified cells in an existing Excel file.

        :param file_path: str, path to the existing Excel file
        :param sheet_name: str, name of the sheet to write data to
        :param data_to_write: dict, key-value pairs where key is the cell reference and value is the value to write
        :param formulas_to_write: dict, key-value pairs where key is the cell reference and value is the formula to write (optional)
        """
        # Load the workbook and select the specified sheet
        wb = load_workbook(file_path)
        if sheet_name not in wb.sheetnames:
            print(f"Sheet name '{sheet_name}' does not exist in the workbook. Adding new sheet.")
            ws = wb.create_sheet(title=sheet_name, index=0)
        else:
            ws = wb[sheet_name]

        # Write the data to the cells
        for cell, value in data_to_write.items():
            ws[cell] = value

        # Write the formulas to the cells
        if formulas_to_write:
            for cell, formula in formulas_to_write.items():
                ws[cell] = f"={formula}"

        # Save the workbook
        wb.save(file_path)
        print(f"Data and formulas written to {file_path} in sheet '{sheet_name}'.")

    @staticmethod
    def add_scatter_chart_to_sheet(file_path, sheet_name, x_col, y_col, x_min_row, x_max_row, y_min_row, y_max_row):
        # Load the workbook and the specified sheet
        wb = load_workbook(file_path)
        ws = wb[sheet_name]

        # Create a scatter chart object
        chart = ScatterChart()
        chart.title = "Scatter Chart"
        chart.x_axis.title = 'X Data'
        chart.y_axis.title = 'Y Data'
        chart.style = 13

        # Create data references for the chart
        x_data = Reference(ws, min_col=x_col, min_row=x_min_row, max_row=x_max_row)
        y_data = Reference(ws, min_col=y_col, min_row=y_min_row, max_row=y_max_row)

        # Create a series and append it to the chart
        series = Series(y_data, x_data)
        chart.series.append(series)

        # Add the chart to the sheet
        ws.add_chart(chart, f"{get_column_letter(x_col + 2)}{y_min_row}")  # Place the chart right next to the data

        # Save the workbook
        wb.save(file_path)
        print(f"Chart added to {sheet_name}.")

    @staticmethod
    def add_multiple_scatter_charts_to_sheet(file_path, sheet_name, x_cols, y_cols, x_min_rows, x_max_rows, y_min_rows, y_max_rows):
        # Validate input lengths
        if not (len(x_cols) == len(y_cols) == len(x_min_rows) == len(x_max_rows) == len(y_min_rows) == len(y_max_rows)):
            raise ValueError("All input lists must have the same length.")

        # Load the workbook and the specified sheet
        wb = load_workbook(file_path)
        ws = wb[sheet_name]

        # Create a scatter chart object
        chart = ScatterChart()
        chart.title = "Combined Scatter Chart"
        chart.x_axis.title = 'X Data'
        chart.y_axis.title = 'Y Data'
        chart.style = 13

        # Iterate over the datasets
        for i, (x_col, y_col, x_min_row, x_max_row, y_min_row, y_max_row) in enumerate(
                zip(x_cols, y_cols, x_min_rows, x_max_rows, y_min_rows, y_max_rows)):
            # Create data references for the chart
            x_data = Reference(ws, min_col=x_col, min_row=x_min_row, max_row=x_max_row)
            y_data = Reference(ws, min_col=y_col, min_row=y_min_row, max_row=y_max_row)

            # Create a series and append it to the chart
            series = Series(y_data, x_data, title=f"Series {i + 1}")
            chart.series.append(series)

        # Add the chart to the sheet, placing it next to the last dataset
        ws.add_chart(chart, f"{get_column_letter(x_cols[-1] + 2)}{y_min_rows[-1]}")

        # Save the workbook
        wb.save(file_path)
        print(f"Chart with multiple series added to {sheet_name}.")


class DataFrameToExcel:

    def __init__(self, excel_path):
        self._exel_path = excel_path

    def write(self):
        pass
